"""Excel (XLSX) 변환기"""

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from .base import Asset, ConversionResult, Converter

logger = logging.getLogger(__name__)


class XlsxConverter(Converter):
    """Excel XLSX 파일을 Markdown으로 변환하는 변환기"""

    supported_extensions = [".xlsx"]

    def convert(self, file_path: Path) -> ConversionResult:
        """
        XLSX 파일을 Markdown으로 변환

        Args:
            file_path: XLSX 파일 경로

        Returns:
            ConversionResult: 변환된 Markdown
        """
        logger.info(f"XLSX 변환 시작: {file_path}")

        markdown_parts: list[str] = []

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        total_sheets = len(sheet_names)

        for sheet_idx, sheet_name in enumerate(sheet_names, start=1):
            ws = wb[sheet_name]

            # 시트 헤더
            if total_sheets > 1:
                markdown_parts.append(f"## {sheet_name}")

            # 시트 내용을 테이블로 변환
            table_md = self._process_sheet(ws)
            if table_md:
                markdown_parts.append(table_md)
            else:
                markdown_parts.append("*(빈 시트)*")

            # 시트 간 구분선 (마지막 시트 제외)
            if sheet_idx < total_sheets:
                markdown_parts.append("\n---\n")

        wb.close()

        markdown = "\n\n".join(markdown_parts)
        markdown = self._cleanup_markdown(markdown)

        logger.info(f"XLSX 변환 완료: {total_sheets}개 시트")

        return ConversionResult(markdown=markdown, assets=[])

    def _process_sheet(self, ws: Worksheet) -> str:
        """워크시트를 Markdown 테이블로 변환"""
        rows: list[list[str]] = []

        # 데이터가 있는 범위 확인
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0 or max_col == 0:
            return ""

        # 모든 행 읽기
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            cells = []
            for cell in row:
                cells.append(self._get_cell_value(cell))
            rows.append(cells)

        # 빈 행 제거 (앞뒤)
        rows = self._trim_empty_rows(rows)

        if not rows:
            return ""

        # 빈 열 제거
        rows = self._trim_empty_cols(rows)

        if not rows or not rows[0]:
            return ""

        # Markdown 테이블 생성
        return self._rows_to_markdown_table(rows)

    def _get_cell_value(self, cell: Cell) -> str:
        """셀 값을 문자열로 변환"""
        if cell.value is None:
            return ""

        value = cell.value

        # 숫자 포맷 처리
        if isinstance(value, float):
            # 정수로 표현 가능하면 정수로
            if value == int(value):
                return str(int(value))
            return str(value)

        if isinstance(value, bool):
            return "Yes" if value else "No"

        # 문자열 변환
        text = str(value).strip()

        # Markdown 테이블에서 문제가 될 수 있는 문자 이스케이프
        text = text.replace("|", "\\|")
        text = text.replace("\n", " ")
        text = text.replace("\r", "")

        return text

    def _trim_empty_rows(self, rows: list[list[str]]) -> list[list[str]]:
        """앞뒤의 빈 행 제거"""
        # 앞쪽 빈 행 제거
        while rows and all(not cell for cell in rows[0]):
            rows.pop(0)

        # 뒤쪽 빈 행 제거
        while rows and all(not cell for cell in rows[-1]):
            rows.pop()

        return rows

    def _trim_empty_cols(self, rows: list[list[str]]) -> list[list[str]]:
        """앞뒤의 빈 열 제거"""
        if not rows:
            return rows

        num_cols = len(rows[0])

        # 빈 열 인덱스 찾기 (앞에서부터)
        start_col = 0
        for col_idx in range(num_cols):
            if any(row[col_idx] for row in rows if col_idx < len(row)):
                start_col = col_idx
                break

        # 빈 열 인덱스 찾기 (뒤에서부터)
        end_col = num_cols
        for col_idx in range(num_cols - 1, -1, -1):
            if any(row[col_idx] for row in rows if col_idx < len(row)):
                end_col = col_idx + 1
                break

        # 열 범위 적용
        return [row[start_col:end_col] for row in rows]

    def _rows_to_markdown_table(self, rows: list[list[str]]) -> str:
        """행 데이터를 Markdown 테이블로 변환"""
        if not rows:
            return ""

        num_cols = max(len(row) for row in rows)

        # 모든 행의 열 수 맞추기
        normalized_rows = []
        for row in rows:
            normalized = row + [""] * (num_cols - len(row))
            normalized_rows.append(normalized)

        md_lines: list[str] = []

        # 헤더 행
        header = normalized_rows[0]
        md_lines.append("| " + " | ".join(header) + " |")

        # 구분선
        md_lines.append("| " + " | ".join(["---"] * num_cols) + " |")

        # 데이터 행
        for row in normalized_rows[1:]:
            md_lines.append("| " + " | ".join(row) + " |")

        return "\n".join(md_lines)

    def _cleanup_markdown(self, text: str) -> str:
        """Markdown 텍스트 정리"""
        lines = text.split("\n")
        cleaned_lines: list[str] = []
        empty_count = 0

        for line in lines:
            if not line.strip():
                empty_count += 1
                if empty_count <= 2:
                    cleaned_lines.append(line)
            else:
                empty_count = 0
                cleaned_lines.append(line)

        return "\n".join(cleaned_lines).strip()
